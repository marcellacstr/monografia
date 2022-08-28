from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from nltk.tokenize import RegexpTokenizer
from time import sleep
import openpyxl
import os
import re

'''até agora esse código faz oq? 

entra no site do stj e busca alguma coisa
coleta dados do primeiro acórdão do STJ resultado da pesquisa 
quais dados? classe processual, número do processo, turma, relator e data do julgamento 
agora também acha se, no acórdão final, tem "negar provimento" ou "dar provimento"
coloca esses dados numa tabela '''

# WEB SCRAPPING - Achando os dados

# abrindo o navegador. Entre '' tá o endereço do chrome webdriver no meu pc
# antes era assim navegador = webdriver.Chrome('/Users/Marcella/PycharmProjects/webscrapper101/chromedriver') mas ficou deprecado

s = Service('/Users/Marcella/PycharmProjects/webscrapper101/chromedriver')
navegador = webdriver.Chrome(service=s)

# A janela do Chrome entra na URL indicada
navegador.get('https://scon.stj.jus.br/SCON/')

# Dentro da URL indicada o programa acha um elemento pelo caminho CSS indicado. Isso fica guardado numa variável
caixaPesquisa = navegador.find_element(By.CSS_SELECTOR, '#pesquisaLivre')

# Manda coisas a serem digitadas na Caixa de Pesquisa que ele achou pelo caminho CSS
caixaPesquisa.send_keys(
    '("ARESP".CLAS. OU "RESP".CLAS.)("ABSOLVICAO" OU "DESCLASSIFICACAO") E ("DROGAS" OU "ENTORPECENTES" OU "11.343")')  # TERMO DE PESQUISA

# aperta enter na pesquisa
caixaPesquisa.submit()

sleep(2)

# coloca para 50 resultados a serem mostrados na página

selecione = Select(navegador.find_element(By.XPATH, '//*[@id="qtdDocsPagina"]'))

selecione.select_by_visible_text('50')

sleep(2)

# ABRINDO O EXCEL

# muda o diretório que o pycharm está trabalhando
os.chdir('/Users/Marcella/Downloads')

# abre (ou carrega) a planilha no excel
planilha = openpyxl.load_workbook('TCCv1.xlsx')

# pega a primeira folha da planilha. (aquelas abas debaixo). O caminho de antes era get sheet by name
folha1 = planilha['Planilha1']

# acha o número de resultados da página

ndeAcordaos = navegador.find_element(By.CSS_SELECTOR,
                                     '#corpopaginajurisprudencia > div.navegacaoDocumento > div.documentoWrapper > div.row.py-1.mb-2.infoPesquisa > div.col > div > span.numDocs')

ndeAcordaosInt = int(ndeAcordaos.text.split()[0])  # acha o número de resultados e transforma de str para int para poder fazer conta

ndePaginas = (ndeAcordaosInt // 50) + 1  # acha o número de páginas, considerando cada página com 50 acórdãos

ndaUltimaPagina = ndeAcordaosInt % 50  # faz divisão inteira por 50 e acha o resto. acha o número de acórdãos da ultima pagina, que vai ter menos de 50

n = 0
ndeResultados = 51  # pq cada página tem 50 resultados normalmente, mas na última página tem menos

for n in range(0, ndePaginas):
    for i in range(1, ndeResultados):
        # procura informações processuais no iºAcórdão
        infoProcessual = navegador.find_element(By.XPATH,
                                                '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[3]/div[1]/div/div[2]' % (
                                                    i))
        # Cria uma lista com as informações processuais, cada elemento da lista dividido pela quebra de linha.
        # EX: de infoProcessual.text
        # '''AgRg no HC 729048 / SP
        # AGRAVO REGIMENTAL NO HABEAS CORPUS
        # 2022/0072411-0'''

        listaInfoProc = infoProcessual.text.splitlines()

        # o que cada elemento da lista é?
        # print(listaInfoProc[0]) #nº do processo Ex: AgRg no HC 729048 / SP
        # print(listaInfoProc[1]) #Classe Processual Ex: AGRAVO REGIMENTAL NO HABEAS CORPUS

        # achando o Ministro Relator
        relatoria = navegador.find_element(By.XPATH,
                                           '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[4]/div[1]/div/div[2]' % (
                                               i))

        # pega o texto da variável relatoria (que ele achou em cima).
        # Depois dá um split em todas as palavras. Depois pega uma lista com as 3 primeiras palavras e junta (join) de novo com um espaço
        valorRelator = ' '.join(relatoria.text.split()[0:3])

        # achando a turma
        turma = navegador.find_element(By.XPATH,
                                       '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[4]/div[2]/div/div[2]' % (
                                           i))
        valorTurma = ''.join(turma.text.split()[0])

        # achando a data do julgamento
        dataJulgamento = navegador.find_element(By.XPATH,
                                                '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[4]/div[3]/div/div[2]' % (
                                                    i))

        # achando o resultado da decisão (deram ou não provimento)
        acordao = navegador.find_element(By.XPATH,
                                         '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[6]/div/div/div[2]/p' % (
                                             i))
        resultado = re.search(r"negar.*provimento|dar.*provimento",
                              acordao.text)  # no TEXTO do elemento que eu achei em cima procure a regex

        # achando o fundamento...

        # primeiro acha a ementa
        ementacompleta = navegador.find_element(By.XPATH,
                                                '//*[@id="corpopaginajurisprudencia"]/div[4]/div[2]/div[2]/div[%i]/div[5]/div/div/div[2]' % (
                                                    i))

        # deixa a ementa toda minuscula e como token e sem sinais de pontuação (., :)
        ementaToken = RegexpTokenizer(r'\w+').tokenize(ementacompleta.text.lower())
        # ementa de volta pra uma string, mas toda minuscula e sem sinais gráficos
        ementaLimpa = ' '.join(ementaToken)
        # fundamento
        fundamento = re.search(r"(?<=s.{1}mula\s).*(?=\sstj)",
                               ementaLimpa)  # procura palavras entre súmula ou sumula (sem acento) e stj

        # COLOCANDO DADOS NO EXCEL

        # definindo a celula que vai a classe processual
        celulaClasseProcessual = folha1.cell(row=((i + 1) + n*50), column=1)

        # na Folha1, na célula A2, coloca a Classe Processual (que tá na lista que foi criada a partir do elemento achado por xpath)
        celulaClasseProcessual.value = listaInfoProc[1]

        # definindo a célula nºdo processo
        celulaNProcesso = folha1.cell(row=((i + 1) + n*50), column=2)

        # na Folha1, na célula B2, coloca o nºdo Processo (que tá na lista)
        celulaNProcesso.value = listaInfoProc[0]

        # definindo a célula que vai a relatoria
        celulaRelatoria = folha1.cell(row=((i + 1) + n*50), column=3)

        # na Folha1, na célula C2, coloca a Relatoria
        celulaRelatoria.value = valorRelator

        # definindo a célula que vai a Turma
        celulaTurma = folha1.cell(row=((i + 1) + n*50), column=4)

        # na Folha1, na célula D2, coloca a TURMA
        celulaTurma.value = valorTurma

        # definindo a célula que vai a Data de Julgamento
        celulaData = folha1.cell(row=((i + 1) + n*50), column=5)

        # na Folha1, na célula E2, coloca a DATA de julgamento
        celulaData.value = dataJulgamento.text

        # definindo a célula que vai o fundamento
        celulaFundamento = folha1.cell(row=((i + 1) + n*50), column=7)

        # na Folha1, na célula G2, coloca o Fundamento (quando?)
        if fundamento == None:
            celulaFundamento.value = 'None'
        else:
            if '7' in fundamento.group():
                celulaFundamento.value = 'sumula 7'  # coloca 7 na célula
            else:
                celulaFundamento.value = 'outra sumula'

        # definindo a célula que vai o resultado
        celulaResultado = folha1.cell(row=((i + 1) + n*50), column=8)

        # na Folha1, na célula H2, coloca o Resultado (quando?) construir um IF
        if resultado == None:
            celulaResultado.value = 'None'
        else:
            celulaResultado.value = resultado.group()
        i += 1
    i = 1
    try:
        navegador.find_element(By.CSS_SELECTOR, '#navegacao > div:nth-child(2) > a.iconeProximaPagina')
    except NoSuchElementException:
        break
    # acha o ícone para ir para a próxima página
    proxPagina = navegador.find_element(By.CSS_SELECTOR, '#navegacao > div:nth-child(2) > a.iconeProximaPagina')
    # clica no ícone para ir para a próxima página
    proxPagina.click()
    sleep(3)
    n += 1
    if n == ndePaginas - 1:
        ndeResultados = ndaUltimaPagina + 1

# isso aqui efetivamente salva na planilha o valor novo atribuído
planilha.save('TCCv1.xlsx')
